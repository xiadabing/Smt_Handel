#!/usr/bin/python
# -*-coding:UTF-8 -*-

# ========================
# @Time  : 2021-07-12
# @Author: Sunny
# ========================

import os
import re
from datetime import datetime

from openpyxl import load_workbook


class HandleExecl(object):
    """
    定义处理excel的类
    """

    def __init__(self):
        self.filename = "order.xlsx"
        self.sheetname = "Sheet1"
        # self.wb = load_workbook(self.filename)  #打开文件
        # self.ws = self.wb[self.sheetname] #定位表单

    def write_result(self, date, shop, number, amount, country):
        other_wb = load_workbook(self.filename)
        other_ws = other_wb[self.sheetname]
        max_row = other_ws.max_row
        other_ws.cell(row=max_row + 1, column=1, value=date)
        other_ws.cell(row=max_row + 1, column=2, value=shop)
        other_ws.cell(row=max_row + 1, column=3, value=number)
        other_ws.cell(row=max_row + 1, column=6, value=amount)
        other_ws.cell(row=max_row + 1, column=7, value=country)
        other_wb.save(self.filename)


def data_hanle(a):
    one_list = []
    key = ''
    # 过滤列表中的邮箱，且返回一个新列表
    for j in a:
        if re.match(r'^[0-9a-zA-Z_]{0,19}@[0-9a-zA-Z]{1,13}\.[com,cn,net]{1,3}$', j):
            a.remove(j)

    # 组成{店铺：订单}嵌套字典的列表
    for i in range(0, len(a)):
        obj = re.match(r"^[a-zA-z].*", a[i])
        if obj:
            key = a[i]
        else:
            value = a[i]
        if i > 0 and obj is None:
            one_list.append({key: value})
    return one_list


# # #配置路径
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
#
# # #读取文件数据
with open(BASE_DIR+"/test.txt",'r') as file:
    context = file.read()
a = context.split('\n')
if a[-1] == '':
    a.pop()

# 数据类型转化
new_date = data_hanle(a)

# 获取当前系统时间
# date = datetime.strftime(datetime.now(), "%Y-%m-%d")

# 写入文档
for item in new_date:
    value = list(item.values())[0].split(' ')
    HandleExecl().write_result(date="2021-07-13", shop=list(item.keys())[0], number=value[0], amount=value[1],
                               country=value[2])